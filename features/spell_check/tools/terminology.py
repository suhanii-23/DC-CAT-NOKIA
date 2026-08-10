"""Tool 1 — Nokia terminology protection.

Determines whether a word is an approved Nokia/domain term that must never be
reported as a spelling error. Two independent protections:

1. An exact-match allow-list loaded at runtime from the DC CAT Acronym/Parameter
   Excel (a column named "Acronym"). Loaded into a set for O(1) lookup.
2. An identifier-shape rule: tokens that are ALL-CAPS, contain digits or
   underscores, or have internal capitals (gNodeB) are structural identifiers,
   not prose words, and are cleared without needing to be in the list.

Exact match is deliberate: "gNodeB" is *similar* to "eNodeB", and similarity is
not equality. A lookup is faster, exact, and auditable — embeddings would be
wrong here.

The allow-list is optional at import time: if no Excel is provided, only the
shape rule applies. This keeps the feature runnable offline and in tests, while
letting the real Nokia vocabulary drop in without code changes.
"""
from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Iterable, Optional

logger = logging.getLogger(__name__)

# Internal capital (gNodeB), any digit (X2), or underscore (airscale_rnc).
_IDENTIFIER_RE = re.compile(r"[a-z][A-Z]|\d|_")

# Common Nokia/telecom terms guaranteed protected even with no Excel present.
# The authoritative list is the runtime Excel; this is a safety net so the
# feature behaves correctly out of the box and in offline tests.
_DEFAULT_TERMS: frozenset[str] = frozenset(
    {
        "gnodeb", "enodeb", "mocn", "rsrp", "rsrq", "sinr",
        "airscale_rnc", "nodeb", "ranslicing", "oss", "bts",
    }
)

_TERMS_COLUMN_ALIASES = ("Acronym", "Abbreviated Name", "Term", "Parameter")


class TerminologyTool:
    """Approved-term membership testing."""

    name = "check_nokia_term"

    def __init__(self, terms: Optional[Iterable[str]] = None) -> None:
        base = set(_DEFAULT_TERMS)
        if terms:
            base |= {t.strip().lower() for t in terms if t and t.strip()}
        self._terms = base

    @classmethod
    def from_excel(cls, path: str | Path) -> "TerminologyTool":
        """Build from a Nokia Acronym/Parameter Excel.

        Accepts any column named ``Acronym`` (per the DC CAT user guide) or a
        small set of aliases. Falls back to defaults if the file is missing or
        unreadable — the tool never crashes the run over a bad list.
        """
        path = Path(path)
        if not path.is_file():
            logger.warning("Terminology Excel not found (%s); using defaults", path)
            return cls()
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            header = [str(c).strip() if c is not None else "" for c in next(rows)]
            col = _pick_column(header)
            if col is None:
                logger.warning("No Acronym-like column in %s; using defaults", path.name)
                return cls()
            idx = header.index(col)
            terms = [str(r[idx]) for r in rows if idx < len(r) and r[idx] is not None]
            logger.info("Loaded %d Nokia terms from %s", len(terms), path.name)
            return cls(terms)
        except Exception as exc:  # never let a bad Excel abort the run
            logger.warning("Could not read terminology Excel %s: %s", path, exc)
            return cls()

    def is_protected(self, word: str) -> bool:
        """True if ``word`` is an approved term or identifier-shaped."""
        if not word:
            return False
        if _IDENTIFIER_RE.search(word):
            return True
        if word.isupper() and len(word) > 1:
            return True
        return word.strip().lower() in self._terms

    def __contains__(self, word: str) -> bool:
        return self.is_protected(word)

    @property
    def size(self) -> int:
        return len(self._terms)


def _pick_column(header: list[str]) -> Optional[str]:
    lowered = {h.lower(): h for h in header}
    for alias in _TERMS_COLUMN_ALIASES:
        if alias.lower() in lowered:
            return lowered[alias.lower()]
    return None
