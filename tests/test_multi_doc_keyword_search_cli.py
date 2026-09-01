"""Tests for features/multi_doc_keyword_search/cli.py — the search CLI.

Focus is the Excel-report policy: a search covering several documents
writes a report without being asked, a single-document one does not, and
the flags override both ways.

Additive: app/cli.py and its behaviour are not involved here — this CLI
lives inside the feature folder and is a separate command.
"""
from __future__ import annotations

import importlib
import os

import pytest

from openpyxl import load_workbook

cli = importlib.import_module("features.multi_doc_keyword_search.cli")


def _pdf(path, text: str) -> str:
    """Write a one-page PDF containing `text`, return its path."""
    fitz = pytest.importorskip("pymupdf")
    doc = fitz.open()
    page = doc.new_page()
    page.insert_text((72, 72), text, fontsize=12)
    doc.save(str(path))
    doc.close()
    return str(path)


def _xlsx_files(folder) -> list[str]:
    return sorted(name for name in os.listdir(folder) if name.endswith(".xlsx"))


# --- automatic report for multiple documents ---------------------------


def test_multiple_documents_write_an_excel_report_without_being_asked(
    tmp_path, monkeypatch, capsys
):
    _pdf(tmp_path / "a.pdf", "Authentication is required.")
    _pdf(tmp_path / "b.pdf", "Authentication retries are logged.")
    monkeypatch.chdir(tmp_path)

    assert cli.main([str(tmp_path), "--keyword", "authentication"]) == 0

    written = _xlsx_files(tmp_path)
    assert written == ["keyword_matches_authentication.xlsx"]
    assert "Excel report written to" in capsys.readouterr().out


def test_the_automatic_report_contains_every_match(tmp_path, monkeypatch):
    _pdf(tmp_path / "a.pdf", "Authentication is required. Retry authentication now.")
    _pdf(tmp_path / "b.pdf", "The AUTHENTICATION token expires.")
    monkeypatch.chdir(tmp_path)

    cli.main([str(tmp_path), "--keyword", "authentication"])

    workbook = load_workbook(tmp_path / "keyword_matches_authentication.xlsx")
    sheet = workbook["multi_doc_keyword_search"]
    rows = list(sheet.iter_rows(values_only=True))
    header, data = rows[0], rows[1:]

    assert len(data) == 3  # one row per occurrence, across both documents
    assert "keyword" in header and "context" in header and "document" in header
    documents = {row[header.index("document")] for row in data}
    assert len(documents) == 2


def test_a_report_is_written_even_when_multiple_documents_have_no_matches(
    tmp_path, monkeypatch
):
    _pdf(tmp_path / "a.pdf", "Configuration overview.")
    _pdf(tmp_path / "b.pdf", "Topology diagram.")
    monkeypatch.chdir(tmp_path)

    assert cli.main([str(tmp_path), "--keyword", "authentication"]) == 0
    assert _xlsx_files(tmp_path) == ["keyword_matches_authentication.xlsx"]


# --- single document, and the overrides --------------------------------


def test_a_single_document_writes_no_report_by_default(tmp_path, monkeypatch):
    path = _pdf(tmp_path / "a.pdf", "Authentication is required.")
    monkeypatch.chdir(tmp_path)

    assert cli.main([path, "--keyword", "authentication"]) == 0
    assert _xlsx_files(tmp_path) == []


def test_excel_flag_forces_a_report_for_a_single_document(tmp_path, monkeypatch):
    path = _pdf(tmp_path / "a.pdf", "Authentication is required.")
    monkeypatch.chdir(tmp_path)

    assert cli.main([path, "--keyword", "authentication", "--excel", "chosen.xlsx"]) == 0
    assert _xlsx_files(tmp_path) == ["chosen.xlsx"]


def test_excel_flag_overrides_the_automatic_path(tmp_path, monkeypatch):
    _pdf(tmp_path / "a.pdf", "Authentication is required.")
    _pdf(tmp_path / "b.pdf", "Authentication retries.")
    monkeypatch.chdir(tmp_path)

    cli.main([str(tmp_path), "--keyword", "authentication", "--excel", "chosen.xlsx"])

    assert _xlsx_files(tmp_path) == ["chosen.xlsx"]


def test_no_excel_suppresses_the_automatic_report(tmp_path, monkeypatch):
    _pdf(tmp_path / "a.pdf", "Authentication is required.")
    _pdf(tmp_path / "b.pdf", "Authentication retries.")
    monkeypatch.chdir(tmp_path)

    assert cli.main([str(tmp_path), "--keyword", "authentication", "--no-excel"]) == 0
    assert _xlsx_files(tmp_path) == []


def test_excel_and_no_excel_are_mutually_exclusive(tmp_path):
    path = _pdf(tmp_path / "a.pdf", "Authentication.")
    with pytest.raises(SystemExit):
        cli.main([path, "--keyword", "authentication", "--excel", "x.xlsx", "--no-excel"])


# --- report naming -----------------------------------------------------


@pytest.mark.parametrize(
    "keyword, expected",
    [
        ("authentication", "keyword_matches_authentication.xlsx"),
        ("MOCN", "keyword_matches_mocn.xlsx"),
        ("access control", "keyword_matches_access_control.xlsx"),
        ("ERR#01", "keyword_matches_err_01.xlsx"),
        ("../etc", "keyword_matches_etc.xlsx"),  # never escapes the directory
        ("###", "keyword_matches_search.xlsx"),  # nothing usable left
    ],
)
def test_default_report_path_is_a_safe_filename(keyword, expected):
    assert cli._default_report_path(keyword) == expected


# --- failure path ------------------------------------------------------


def test_an_empty_keyword_writes_no_report_and_fails(tmp_path, monkeypatch, capsys):
    _pdf(tmp_path / "a.pdf", "Authentication is required.")
    _pdf(tmp_path / "b.pdf", "Authentication retries.")
    monkeypatch.chdir(tmp_path)

    assert cli.main([str(tmp_path), "--keyword", "   "]) == 2
    assert _xlsx_files(tmp_path) == []
    assert "Search failed" in capsys.readouterr().err
