"""Integration tests: the feature as wired into app/cli.py and app/mcp_server.py.

Three things are checked here that the unit tests cannot see:

  1. `python -m app.cli ... --features multi_doc_keyword_search` runs the
     pipeline exactly once over the whole document set, not once per
     document.
  2. `--features all` includes it when a --query is given, and skips it
     (rather than failing) when there is no keyword to search for, so the
     documented no---query command keeps working.
  3. The MCP server registers the new tool alongside the existing three,
     and the existing three are still registered.
"""
from __future__ import annotations

import importlib

import pytest

from openpyxl import load_workbook

cli = importlib.import_module("app.cli")


def _pdf(path, text: str) -> str:
    fitz = pytest.importorskip("pymupdf")
    doc = fitz.open()
    page = doc.new_page()
    page.insert_text((72, 72), text, fontsize=12)
    doc.save(str(path))
    doc.close()
    return str(path)


@pytest.fixture
def corpus(tmp_path):
    """Three documents: 2 matches, 1 match, 0 matches."""
    _pdf(tmp_path / "a.pdf", "Authentication is required. Retry authentication now.")
    _pdf(tmp_path / "b.pdf", "The AUTHENTICATION token expires hourly.")
    _pdf(tmp_path / "c.pdf", "Topology diagram only.")
    return tmp_path


# --- the pipeline is reachable from the main CLI -----------------------


def test_pipeline_runs_when_named_explicitly(corpus, capsys):
    exit_code = cli.main(
        [str(corpus), "--features", "multi_doc_keyword_search", "--query", "authentication"]
    )
    out = capsys.readouterr().out

    assert exit_code == 0
    assert "=== multi_doc_keyword_search (3 document(s)) ===" in out
    assert "[multi_doc_keyword_search] status=ok findings=3" in out


def test_pipeline_runs_once_over_the_whole_set_not_once_per_document(corpus, capsys):
    cli.main(
        [str(corpus), "--features", "multi_doc_keyword_search", "--query", "authentication"]
    )
    out = capsys.readouterr().out

    # One pipeline block, not one per document.
    assert out.count("[multi_doc_keyword_search] status=") == 1
    assert "a.pdf" in out and "b.pdf" in out  # both documents represented


def test_pipeline_output_carries_page_and_context(corpus, capsys):
    cli.main(
        [str(corpus), "--features", "multi_doc_keyword_search", "--query", "authentication"]
    )
    out = capsys.readouterr().out

    assert "(p1)" in out
    assert "Authentication is required" in out


def test_pipeline_writes_into_the_shared_excel_report(corpus, tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    cli.main(
        [
            str(corpus),
            "--features",
            "multi_doc_keyword_search",
            "--query",
            "authentication",
            "--excel",
            "report.xlsx",
        ]
    )

    workbook = load_workbook(tmp_path / "report.xlsx")
    assert "multi_doc_keyword_search" in workbook.sheetnames
    rows = list(workbook["multi_doc_keyword_search"].iter_rows(values_only=True))
    assert len(rows) == 4  # header + one row per occurrence
    assert "context" in rows[0] and "document" in rows[0]


def test_pipeline_can_run_alongside_a_per_document_feature(corpus, capsys):
    exit_code = cli.main(
        [
            str(corpus),
            "--features",
            "broken_links,multi_doc_keyword_search",
            "--query",
            "authentication",
        ]
    )
    out = capsys.readouterr().out

    assert exit_code == 0
    assert "[broken_links] status=" in out
    assert "[multi_doc_keyword_search] status=ok" in out


def test_a_missing_keyword_skips_the_pipeline_without_killing_the_run(corpus, capsys):
    exit_code = cli.main([str(corpus), "--features", "multi_doc_keyword_search"])
    captured = capsys.readouterr()

    assert exit_code == 0
    assert "[multi_doc_keyword_search] status=skipped findings=0" in captured.out
    assert "needs --query" in captured.err


# --- "all" includes the pipeline ---------------------------------------


def test_features_all_runs_the_pipeline_when_a_query_is_given(corpus, capsys):
    cli.main([str(corpus), "--query", "authentication"])  # default is "all"
    out = capsys.readouterr().out

    assert "[multi_doc_keyword_search] status=ok findings=3" in out


def test_features_all_still_runs_every_per_document_feature(corpus, capsys):
    cli.main([str(corpus), "--query", "authentication"])
    out = capsys.readouterr().out

    for feature in cli.FEATURE_MODULES:
        assert f"[{feature.name}] status=" in out


def test_features_all_without_a_query_skips_the_pipeline(corpus, capsys):
    # The documented `python -m app.cli <file> --excel report.xlsx` run.
    exit_code = cli.main([str(corpus)])
    captured = capsys.readouterr()

    assert exit_code == 0
    assert "[multi_doc_keyword_search] status=skipped" in captured.out
    assert "status=failed" not in captured.out
    assert "needs --query" in captured.err


def test_the_corpus_feature_is_not_a_per_document_feature(corpus):
    assert "multi_doc_keyword_search" in cli.ALL_FEATURE_NAMES
    assert "multi_doc_keyword_search" in cli.CORPUS_NAMES
    assert "multi_doc_keyword_search" not in {f.name for f in cli.FEATURE_MODULES}


def test_an_unknown_feature_name_is_still_rejected(corpus):
    with pytest.raises(SystemExit):
        cli.main([str(corpus), "--features", "no_such_feature"])


def test_naming_the_pipeline_does_not_print_a_not_implemented_note(corpus, capsys):
    cli.main(
        [str(corpus), "--features", "multi_doc_keyword_search", "--query", "authentication"]
    )
    assert "not implemented yet" not in capsys.readouterr().err


# --- one corrupt document does not stop the run ------------------------


def test_a_corrupt_document_does_not_stop_the_pipeline_or_the_other_features(
    corpus, capsys
):
    (corpus / "corrupt.pdf").write_bytes(b"this is not a pdf")

    exit_code = cli.main(
        [
            str(corpus),
            "--features",
            "broken_links,multi_doc_keyword_search",
            "--query",
            "authentication",
        ]
    )
    out = capsys.readouterr().out

    assert exit_code == 0
    # The corrupt document is reported as failed, not raised.
    assert "status=failed" in out
    assert "Could not read" in out
    # The readable documents are still processed by both kinds of feature.
    assert out.count("[broken_links] status=ok") == 3
    assert "[multi_doc_keyword_search] status=ok findings=3" in out


def test_both_kinds_of_feature_get_their_own_excel_sheet(corpus, tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    cli.main(
        [
            str(corpus),
            "--features",
            "broken_links,multi_doc_keyword_search",
            "--query",
            "authentication",
            "--excel",
            "mixed.xlsx",
        ]
    )

    workbook = load_workbook(tmp_path / "mixed.xlsx")
    assert "Summary" in workbook.sheetnames
    assert "multi_doc_keyword_search" in workbook.sheetnames
    assert any(name.startswith("broken_links") for name in workbook.sheetnames)
    # The pipeline contributed exactly one FeatureResult, so exactly one sheet.
    assert sum(n.startswith("multi_doc_keyword_search") for n in workbook.sheetnames) == 1


# --- LangGraph orchestration -------------------------------------------


def _graph():
    """The agent graph module, or a skip when langgraph isn't installed.

    app/cli.py treats langgraph as optional and falls back to calling the
    services directly, so the suite has to stay green either way.
    """
    return pytest.importorskip(
        "app.agent.graph", reason="langgraph not installed; CLI uses the fallback"
    )


def test_corpus_graph_runs_the_feature_once_over_the_whole_corpus(corpus):
    graph = _graph()

    state = graph.corpus_graph.invoke(
        {
            "paths": [str(corpus)],
            "options": {"query": "authentication"},
            "selected_features": ["multi_doc_keyword_search"],
        }
    )

    assert len(state["results"]) == 1  # one FeatureResult for the whole corpus
    result = state["results"][0]
    assert result.feature == "multi_doc_keyword_search"
    assert result.status == "ok"
    assert result.meta["documents_searched"] == 3
    assert result.meta["total_matches"] == 3


def test_corpus_graph_skips_when_there_is_no_keyword(corpus):
    graph = _graph()

    state = graph.corpus_graph.invoke(
        {
            "paths": [str(corpus)],
            "options": {"query": None},
            "selected_features": ["multi_doc_keyword_search"],
        }
    )

    assert state["results"][0].status == "skipped"


def test_the_per_document_graph_does_not_know_the_corpus_feature(corpus):
    graph = _graph()

    # The corpus feature must not be reachable from the per-document graph:
    # that is exactly how it would end up running once per document.
    assert "multi_doc_keyword_search" not in graph.route_selected_features(
        {"selected_features": ["multi_doc_keyword_search"]}
    )
    assert graph.route_selected_corpus_features(
        {"selected_features": ["multi_doc_keyword_search"]}
    ) == ["multi_doc_keyword_search"]


def test_the_graph_delegates_and_holds_no_matching_logic():
    import inspect

    graph = _graph()
    source = inspect.getsource(graph)

    assert "re.compile" not in source
    assert "finditer" not in source
    # It reaches the feature only through the service's public entry point.
    assert "multi_doc_keyword_search_service.search(" in source


def test_the_direct_fallback_matches_the_graph(corpus):
    """The no-langgraph path must produce the same result as the graph."""
    results = cli._run_corpus_features_directly(
        [str(corpus)], {"query": "authentication"}, cli.CORPUS_MODULES
    )

    assert len(results) == 1
    assert results[0].status == "ok"
    assert results[0].meta["total_matches"] == 3


def test_the_direct_fallback_skips_without_a_keyword(corpus):
    results = cli._run_corpus_features_directly(
        [str(corpus)], {"query": None}, cli.CORPUS_MODULES
    )

    assert results[0].status == "skipped"


# --- the MCP server manages the feature --------------------------------


def test_mcp_server_registers_the_new_tool_and_keeps_the_existing_ones():
    mcp_server = pytest.importorskip("app.mcp_server")

    for tool in (
        "check_broken_links",
        "search_document",
        "check_spelling",
        "search_documents_for_keyword",
    ):
        assert hasattr(mcp_server, tool), f"{tool} is not registered"


def test_mcp_tool_returns_the_aggregated_payload(corpus):
    mcp_server = pytest.importorskip("app.mcp_server")

    payload = mcp_server.search_documents_for_keyword([str(corpus)], "authentication")

    assert payload["status"] == "ok"
    assert payload["keyword"] == "authentication"
    assert payload["documents_searched"] == 3
    assert payload["documents_with_matches"] == 2
    assert payload["total_matches"] == 3


def test_mcp_tool_reports_an_empty_keyword_as_a_failure(corpus):
    mcp_server = pytest.importorskip("app.mcp_server")

    payload = mcp_server.search_documents_for_keyword([str(corpus)], "  ")

    assert payload["status"] == "failed"
    assert "empty" in payload["error"]
